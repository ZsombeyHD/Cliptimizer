import sqlite3
import tkinter as tk
from tkinter import Toplevel, PhotoImage, OptionMenu, StringVar, messagebox, Spinbox
from fpdf import FPDF
import openpyxl
from openpyxl.styles import Font
import os


class PlanCreatorPage(tk.Frame):
    """A tervezési terv(ek) létrehozására szolgáló oldal."""

    def __init__(self, parent, *args, **kwargs):
        tk.Frame.__init__(self, parent, *args, **kwargs)
        self.configure(bg='white')

        # A termékek tárolására való tömb inicializálása
        self.product_entries = []

        # Az adatbázis kapcsolat
        self.conn = sqlite3.connect('cliptimizer.db')
        self.cursor = self.conn.cursor()

        # Products táblából adat lekérése
        self.cursor.execute("SELECT name FROM products")
        self.product_names = [row[0] for row in self.cursor.fetchall()]

        # Fő konténer a terv létrehozása és panelek számára
        main_container = tk.Frame(self, bg='white')
        main_container.pack(fill=tk.BOTH, expand=True)

        # Bal oldali konténer a terv paneleknek
        left_container = tk.Frame(main_container, bg='white')
        left_container.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Jobb oldali konténer (itt nincs szükség a függeszték státusz megjelenítésére)
        right_container = tk.Frame(main_container, bg='white')
        right_container.pack(side=tk.RIGHT, padx=20, pady=20, anchor=tk.N)

        # A cím
        label = tk.Label(left_container, text="PRÓBASZERŰ TERVEK LÉTREHOZÁSA", bg='white',
                         font=('Helvetica', 20, 'bold'),
                         padx=20, pady=20)
        label.pack(anchor=tk.N)

        # Ezen az oldalon található képek
        self.add_icon = PhotoImage(file='images/add_plan_resized.png')
        self.eye_icon = PhotoImage(file='images/eye_resized.png')
        self.trash_icon = PhotoImage(file='images/trash_resized.png')
        self.print_icon = PhotoImage(file='images/print_resized.png')
        self.excel_icon = PhotoImage(file='images/file-excel_resized.png')
        self.duplicate_icon = PhotoImage(file='images/duplicate_resized.png')

        # Terv létrehozása gomb
        add_button = tk.Button(left_container, image=self.add_icon, bg='white', bd=0, command=self.create_new_plan)
        add_button.pack(anchor=tk.N, pady=20)

        # Terv panelek konténere
        self.plan_container = tk.Frame(left_container, bg='white')
        self.plan_container.pack(anchor=tk.N, pady=20)

        # Tervek betöltése
        self.load_plans()

    def load_plans(self):
        """Tervek betöltése és panelek megjelenítése."""
        self.cursor.execute("SELECT plan_name, SUM(hangers_needed) FROM plans WHERE is_draft = 1 GROUP BY plan_name")
        plans = self.cursor.fetchall()

        for plan in plans:
            plan_name, hangers_needed = plan
            self.add_plan_panel(plan_name, hangers_needed)

    def create_new_plan(self):
        """Egy terv létrehozása: Új ablak és adatbázis kapcsolódás."""
        new_window = Toplevel(self)
        new_window.title("Terv")
        new_window.geometry("1920x1080")
        new_window.configure(bg='white')  # Ablak teljes háttere fehér

        # Ürítjük a product_entries listát, hogy ne legyen widget probléma
        self.product_entries.clear()

        # Fő konténer
        main_container = tk.Frame(new_window, bg='white')
        main_container.pack(fill=tk.BOTH, expand=True)

        # A cím
        label = tk.Label(main_container, text="Terv létrehozása", font=('Helvetica', 14))
        label.pack(pady=20)

        # A terv neve
        plan_name_entry = tk.Entry(main_container, bg='white')
        plan_name_entry.pack(pady=10)

        # Frame a termékeknek és az új termék gombnak
        product_frame = tk.Frame(main_container, bg='white')
        product_frame.pack(expand=True, fill=tk.BOTH, pady=10)

        # Kezdő termék hozzáadása
        self.add_product_field(product_frame)

        # Új termék hozzáadása gomb
        add_product_button = tk.Button(main_container, image=self.add_icon, bg='white', bd=0,
                                       command=lambda: self.add_product_field(product_frame))
        add_product_button.pack(pady=10)

        # Frame a mentés gombnak
        button_frame = tk.Frame(main_container, bg='white')
        button_frame.pack(side=tk.BOTTOM, fill=tk.X, pady=10)

        # Mentés gomb hozzáadása
        save_button = tk.Button(button_frame, text="Mentés", font=('Helvetica', 14), bg='white',
                                command=lambda: self.save_plan(new_window, plan_name_entry.get()))
        save_button.pack(pady=10)

    def add_product_field(self, window):
        """Termék, mennyiség és opcionális manuális függesztékszám mezők hozzáadása."""
        product_frame = tk.Frame(window)
        product_frame.pack(pady=10)

        selected_product = StringVar(window)
        selected_product.set(self.product_names[0])
        product_menu = OptionMenu(product_frame, selected_product, *self.product_names)
        product_menu.pack(side=tk.LEFT, padx=10)

        amount_entry = Spinbox(product_frame, from_=1, to=1000, width=5)
        amount_entry.pack(side=tk.LEFT, padx=10)

        hanger_label = tk.Label(product_frame, text="Függesztékszám: 0", font=('Helvetica', 12))
        hanger_label.pack(side=tk.LEFT, padx=10)

        manual_hanger_var = tk.BooleanVar()
        manual_hanger_check = tk.Checkbutton(product_frame, text="Manuális függesztékszám",
                                             variable=manual_hanger_var, font=('Helvetica', 12))
        manual_hanger_check.pack(side=tk.LEFT, padx=10)

        manual_hanger_entry = tk.Entry(product_frame, font=('Helvetica', 12), state='disabled', width=5)
        manual_hanger_entry.pack(side=tk.LEFT, padx=10)

        def toggle_hanger_entry():
            if manual_hanger_var.get():
                manual_hanger_entry.config(state='normal')
            else:
                manual_hanger_entry.config(state='disabled')

        manual_hanger_check.config(command=toggle_hanger_entry)

        def update_hanger_label(*args):
            product_name = selected_product.get()
            amount = int(amount_entry.get())
            self.cursor.execute("SELECT items_per_hanger FROM products WHERE name = ?", (product_name,))
            items_per_hanger = self.cursor.fetchone()[0]
            required_hangers = (amount + items_per_hanger - 1) // items_per_hanger
            hanger_label.config(text=f"Függesztékszám: {required_hangers}")

        amount_entry.bind("<KeyRelease>", update_hanger_label)
        selected_product.trace("w", update_hanger_label)

        self.product_entries.append((selected_product, amount_entry, manual_hanger_var, manual_hanger_entry))

    def save_plan(self, window, plan_name):
        """Terv mentése."""
        if not plan_name:
            messagebox.showerror("Hiba", "Terv neve nem lehet üres!")
            return

        self.cursor.execute("SELECT COUNT(*) FROM plans WHERE plan_name = ?", (plan_name,))
        plan_exists = self.cursor.fetchone()[0]
        if plan_exists:
            messagebox.showerror("Hiba", "Már létezik ilyen nevű terv!")
            return

        if self.product_entries:
            products_to_save = []
            try:
                for selected_product, amount_entry, manual_hanger_var, manual_hanger_entry in self.product_entries:
                    product_name = selected_product.get()
                    amount = int(amount_entry.get())
                    self.cursor.execute("SELECT id, items_per_hanger FROM products WHERE name = ?", (product_name,))
                    product_data = self.cursor.fetchone()
                    product_id, items_per_hanger = product_data
                    required_hangers = (amount + items_per_hanger - 1) // items_per_hanger

                    if manual_hanger_var.get():
                        manual_hangers = int(manual_hanger_entry.get())
                        required_hangers = manual_hangers

                    products_to_save.append((plan_name, product_id, amount, required_hangers))

                for plan_name, product_id, amount, required_hangers in products_to_save:
                    self.cursor.execute(
                        "INSERT INTO plans (plan_name, product_ID, amount, hangers_needed, is_draft) VALUES"
                        " (?, ?, ?, ?, ?)",
                        (plan_name, product_id, amount, required_hangers, 1)
                    )
                    self.conn.commit()

                self.add_plan_panel(plan_name, sum([entry[3] for entry in products_to_save]))
                window.destroy()

            except Exception as e:
                messagebox.showerror("Hiba", f"Hiba történt a terv mentésekor: {e}")

    def add_plan_panel(self, plan_name, hangers_needed):
        """Panel hozzáadása a tervhez."""
        self.cursor.execute("""
            SELECT p.total_cycle_time, pl.hangers_needed
            FROM plans pl
            JOIN products p ON pl.product_ID = p.id
            WHERE pl.plan_name = ?
        """, (plan_name,))
        plans = self.cursor.fetchall()

        total_cycle_time = sum(hangers_needed * cycle_time for cycle_time, hangers_needed in plans)
        days, remainder = divmod(total_cycle_time, 86400)
        hours, remainder = divmod(remainder, 3600)
        minutes, seconds = divmod(remainder, 60)
        formatted_cycle_time = f"{days}n:{hours}ó:{minutes}p:{seconds}mp"

        panel = tk.Frame(self.plan_container, bg='white', bd=2, relief='solid')
        panel.pack(pady=5, padx=10, fill=tk.X)

        plan_label = tk.Label(panel, text=f"{plan_name}", bg='white', font=('Helvetica', 12))
        plan_label.pack(side=tk.LEFT, padx=10, pady=5)

        hangers_label = tk.Label(panel,
                                 text=f"Elfoglalt függesztékek: {hangers_needed}, Ciklusidő: {formatted_cycle_time}",
                                 bg='white', font=('Helvetica', 10))
        hangers_label.pack(side=tk.LEFT, padx=10, pady=5)

        # Áthelyezés ikon csak akkor, ha a függesztékszám 70 vagy kevesebb
        if hangers_needed <= 70:
            move_icon = PhotoImage(file='images/move_to_active_resized.png')
            move_button = tk.Button(panel, image=move_icon, bg='white', bd=0, command=lambda:
            self.attempt_move_to_active(plan_name, hangers_needed))
            move_button.image = move_icon  # Kép referencia megőrzése
            move_button.pack(side=tk.RIGHT, padx=10, pady=5)

        view_button = tk.Button(panel, image=self.eye_icon, bg='white', bd=0,
                                command=lambda: self.view_plan(plan_name))
        view_button.pack(side=tk.RIGHT, padx=10, pady=5)

        delete_button = tk.Button(panel, image=self.trash_icon, bg='white', bd=0,
                                  command=lambda: self.confirm_delete(plan_name))
        delete_button.pack(side=tk.RIGHT, padx=10, pady=5)

        print_button = tk.Button(panel, image=self.print_icon, bg='white', bd=0,
                                 command=lambda: self.print_plan(plan_name))
        print_button.pack(side=tk.RIGHT, padx=10, pady=5)

        excel_button = tk.Button(panel, image=self.excel_icon, bg='white', bd=0,
                                 command=lambda: self.export_plan_to_excel(plan_name))
        excel_button.pack(side=tk.RIGHT, padx=10, pady=5)

        duplicate_button = tk.Button(panel, image=self.duplicate_icon, bg='white', bd=0,
                                     command=lambda: self.duplicate_plan(plan_name))
        duplicate_button.pack(side=tk.RIGHT, padx=10, pady=5)

    def attempt_move_to_active(self, plan_name, required_hangers):
        """Próbaszerű terv áthelyezése az éles tervek közé, ha elegendő függeszték van."""
        try:
            # Lekérdezzük az elérhető függesztékek számát az éles tervek közül
            self.cursor.execute("SELECT available FROM hangers")
            available_hangers = self.cursor.fetchone()[0]

            # Ellenőrizzük, hogy elegendő függeszték áll-e rendelkezésre
            if required_hangers <= available_hangers:
                # Az éles tervekbe áthelyezés
                self.cursor.execute("UPDATE plans SET is_draft = 0 WHERE plan_name = ?", (plan_name,))
                self.conn.commit()

                # Függesztékek számának frissítése
                self.cursor.execute("UPDATE hangers SET available = available - ?, occupied = occupied + ?",
                                    (required_hangers, required_hangers))
                self.conn.commit()

                # Üzenet a felhasználónak
                messagebox.showinfo("Terv áthelyezve", f"A terv sikeresen áthelyezve az éles tervek közé:"
                                                       f" {plan_name}")

                # Panel eltávolítása a próbaszerű tervekről
                for widget in self.plan_container.winfo_children():
                    if isinstance(widget, tk.Frame):
                        if widget.winfo_children()[0].cget("text") == plan_name:
                            widget.destroy()

            else:
                # Hibaüzenet, ha nincs elegendő függeszték
                messagebox.showerror("Hiba", f"Nincs elég elérhető függeszték az áthelyezéshez. "
                                             f"Szükséges: {required_hangers} függeszték, Elérhető: {available_hangers} "
                                             f"függeszték")

        except Exception as e:
            messagebox.showerror("Hiba", f"Hiba történt az áthelyezés során: {e}")

    def view_plan(self, plan_name):
        """Terv megjelenítése egy új ablakban."""
        new_window = Toplevel(self)
        new_window.title(f"Terv: {plan_name}")
        new_window.geometry("1920x1080")

        label = tk.Label(new_window, text=f"Terv neve: {plan_name}", font=('Helvetica', 12, 'bold'))
        label.pack(pady=10)

        # A termékek megjelenítése a tervhez
        self.cursor.execute("SELECT product_ID, amount, hangers_needed FROM plans WHERE plan_name = ?",
                            (plan_name,))
        products = self.cursor.fetchall()

        for product_id, amount, hangers_needed in products:
            # Lekérjük az összes szükséges attribútumot
            self.cursor.execute(
                "SELECT name, color, clip_type, items_per_hanger, total_cycle_time, material_per_part "
                "FROM products WHERE id = ?",
                (product_id,)
            )
            product = self.cursor.fetchone()
            product_name, color, clip_type, items_per_hanger, total_cycle_time, material_per_part = product

            # Képletek az összes anyagszükséglet és a függesztékek ciklusidejéhez
            total_material = amount * material_per_part
            total_cycle_for_hangers = hangers_needed * total_cycle_time

            # Szükséges körök száma kiszámítása
            items_per_round = 70 * items_per_hanger
            required_cycles = -(-amount // items_per_round)  # Felfelé kerekítés

            # Frame a termékhez tartozó attribútumok megjelenítésére
            frame = tk.Frame(new_window, bg='white', bd=1, relief='solid')
            frame.pack(pady=5, padx=10, fill=tk.X)

            attributes = [
                f"Termék neve: {product_name}",
                f"Mennyiség: {amount}",
                f"Szín: {color}",
                f"Klipsz típusa: {clip_type}",
                f"Függesztékre rakható: {items_per_hanger} db",
                f"Függesztékenként ciklusidő: {total_cycle_time} mp",
                f"Lefoglalt függesztékszám: {hangers_needed}",
                f"Ennyi függesztékre teljes ciklusidő: {total_cycle_for_hangers} mp",
                f"Szükséges körök száma: {required_cycles}",
                f"Anyagszükséglet / alkatrész: {material_per_part:.2f} g",
                f"Összes anyagszükséglet: {total_material:.2f} g"
            ]

            # Attribútumok hozzáadása a frame-hez
            for attr in attributes:
                label = tk.Label(frame, text=attr, font=('Helvetica', 8), bg='white')
                label.pack(side=tk.LEFT, padx=5, pady=5)

    def confirm_delete(self, plan_name):
        """Terv törlésének megerősítése és törlése."""
        if messagebox.askyesno("Törlés megerősítése", f"Biztos törölni szeretné ezt a tervet : "
                                                      f"'{plan_name}'?"):
            self.delete_plan(plan_name)

    def delete_plan(self, plan_name):
        """Terv törlése az adatbázisból és a panelből."""
        try:
            # Töröljük a tervet
            self.cursor.execute("DELETE FROM plans WHERE plan_name = ?", (plan_name,))
            self.conn.commit()

            # Frissítjük a felhasználói felületet
            for widget in self.plan_container.winfo_children():
                if isinstance(widget, tk.Frame):
                    if widget.winfo_children()[0].cget("text") == plan_name:
                        widget.destroy()

        except Exception as e:
            messagebox.showerror("Hiba", f"Hiba történt a terv törlésekor: {e}")

    def print_plan(self, plan_name):
        """Terv nyomtatása PDF formátumban."""
        try:
            pdf = FPDF(orientation='L', unit='mm', format='A4')
            pdf.add_page()

            # Egyedi font
            pdf.add_font('DejaVu', '', 'fonts/DejaVuSans.ttf', uni=True)
            pdf.add_font('DejaVu', 'B', 'fonts/DejaVuSans-Bold.ttf', uni=True)

            # Táblázat attribútumok
            cell_height = 6  # Smaller cell height
            font_size = 7  # Smaller font size
            pdf.set_font('DejaVu', 'B', font_size)

            # A terv neve
            pdf.cell(0, 8, f"TERV: {plan_name}", ln=True, align='L')  # Align left, no margin

            # A headerek
            headers = ["Kód", "Mennyiség", "Szín", "Klipsz", "Füg-re rakható",
                       "Füg-ként ciklusidő", "Lefoglalt füg.", "Anyagszükséglet (g)",
                       "Ennyi füg-re teljes idő", "Szükséges körök"]
            col_widths = [20, 20, 15, 15, 25, 28, 25, 30, 30, 25]  # Optimized column widths
            for i, header in enumerate(headers):
                pdf.cell(col_widths[i], cell_height, header, border=1, align='C')
            pdf.ln(cell_height)

            # A termékek
            pdf.set_font('DejaVu', '', font_size)
            self.cursor.execute("SELECT product_ID, amount, hangers_needed FROM plans WHERE plan_name = ?",
                                (plan_name,))
            products = self.cursor.fetchall()

            for product_id, amount, hangers_needed in products:
                self.cursor.execute(
                    "SELECT name, color, clip_type, items_per_hanger, total_cycle_time, material_per_part FROM "
                    "products WHERE id = ?",
                    (product_id,)
                )
                product = self.cursor.fetchone()
                product_name, color, clip_type, items_per_hanger, total_cycle_time, material_per_part = product

                # Számítások
                total_material = round(material_per_part * amount, 2)
                total_cycle_for_hangers = hangers_needed * total_cycle_time
                items_per_round = 70 * items_per_hanger
                required_cycles = -(-amount // items_per_round)

                # Sor az adatokkal
                row = [
                    product_name, str(amount), color, clip_type, str(items_per_hanger),
                    str(total_cycle_time), str(hangers_needed), f"{total_material} g", f"{total_cycle_for_hangers} mp",
                    str(required_cycles)
                ]

                for i, item in enumerate(row):
                    pdf.cell(col_widths[i], cell_height, item, border=1, align='C')
                pdf.ln(cell_height)

            # A PDF mentése és megnyitása
            pdf_output_path = f"{plan_name}_terv.pdf"
            pdf.output(pdf_output_path)

            try:
                os.startfile(pdf_output_path, "print")
            except FileNotFoundError:
                messagebox.showerror("Hiba", "Nincs alapértelmezett alkalmazás PDF-hez.")
            messagebox.showinfo("Nyomtatás", "A terv nyomtatása folyamatban.")

        except Exception as e:
            messagebox.showerror("Hiba", f"Hiba történt a terv nyomtatása során: {e}")

    def export_plan_to_excel(self, plan_name):
        """Terv exportálása Excel formátumban."""
        try:
            # Excel létrehozása
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = plan_name

            bold_font = Font(bold=True)
            title_font = Font(size=14, bold=True)

            # Terv neve
            ws.merge_cells("A1:J1")
            ws["A1"] = plan_name.upper()
            ws["A1"].font = title_font

            # Adatfejlécek és új oszlop a teljes anyagszükséglet számára
            headers = ["Termék neve", "Mennyiség", "Szín", "Klipsz típusa", "Függesztékre helyezhető darabszám",
                       "Függesztékenként ciklusidő (sec)", "Lefoglalt függesztékek száma", "Összes anyagszükséglet (g)",
                       "Ennyi függesztékre teljes ciklusidő (sec)"]
            ws.append([""] * len(headers))
            ws.append(headers)
            for col in range(1, len(headers) + 1):
                ws.cell(row=3, column=col).font = bold_font

            # Termékek megjelenítése a tervhez
            self.cursor.execute("SELECT product_ID, amount, hangers_needed FROM plans WHERE plan_name = ?",
                                (plan_name,))
            products = self.cursor.fetchall()

            for product_id, amount, reserved_hangers in products:
                self.cursor.execute(
                    "SELECT name, color, clip_type, items_per_hanger, total_cycle_time, material_per_part FROM "
                    "products WHERE id = ?",
                    (product_id,)
                )
                product = self.cursor.fetchone()
                product_name, color, clip_type, items_per_hanger, total_cycle_time, material_per_part = product

                # Anyagszükséglet és teljes ciklusidő számítása
                total_material = amount * material_per_part
                total_cycle_for_hangers = reserved_hangers * total_cycle_time

                # Adatok hozzáadása Excelhez
                ws.append([product_name, amount, color, clip_type, items_per_hanger, total_cycle_time, reserved_hangers,
                           f"{total_material:.2f}", total_cycle_for_hangers])

            # Excel mentése
            excel_output_path = f"{plan_name}_terv.xlsx"
            wb.save(excel_output_path)

            # Excel megnyitása
            os.startfile(excel_output_path)

            messagebox.showinfo("Exportálás", "A terv sikeresen exportálva Excel formátumba.")

        except Exception as e:
            messagebox.showerror("Hiba", f"Hiba történt a terv Excel exportálása során: {e}")

    def duplicate_plan(self, plan_name):
        """Egy terv duplikálása."""
        try:
            # Lekérjük a tervhez tartozó termékeket és azok adatait
            self.cursor.execute("SELECT product_ID, amount FROM plans WHERE plan_name = ?",
                                (plan_name,))
            products = self.cursor.fetchall()

            # Új ablak megnyitása a duplikált terv szerkesztéséhez
            new_window = Toplevel(self)
            new_window.title(f"Duplikált terv: {plan_name}")
            new_window.geometry("1920x1080")

            # Új terv neve (opcionálisan megváltoztatható)
            new_plan_name_entry = tk.Entry(new_window)
            new_plan_name_entry.insert(0, f"{plan_name}_másolat")
            new_plan_name_entry.pack(pady=10)

            # Termékek és mennyiség mezők inicializálása
            product_frame = tk.Frame(new_window)
            product_frame.pack(expand=True, fill=tk.BOTH, pady=10)

            # Termékek hozzáadása a duplikált tervhez
            self.product_entries = []  # Ürítjük a meglévő termékbejegyzéseket
            for product_id, amount in products:
                # Lekérjük a termék részleteit
                self.cursor.execute(
                    "SELECT name FROM products WHERE id = ?",
                    (product_id,)
                )
                product_name = self.cursor.fetchone()[0]

                # Termék mezők megjelenítése
                self.add_product_field(product_frame)
                self.product_entries[-1][0].set(product_name)  # Termék neve
                self.product_entries[-1][1].delete(0, 'end')
                self.product_entries[-1][1].insert(0, amount)  # Mennyiség

            # + ikon hozzáadása a termékekhez
            add_product_button = tk.Button(new_window, image=self.add_icon, bg='white', bd=0,
                                           command=lambda: self.add_product_field(product_frame))
            add_product_button.pack(pady=10)

            # Mentés gomb hozzáadása a duplikált terv mentéséhez
            save_button = tk.Button(new_window, text="Mentés", font=('Helvetica', 14),
                                    command=lambda: self.save_plan(new_window, new_plan_name_entry.get()))
            save_button.pack(pady=10)

        except Exception as e:
            messagebox.showerror("Hiba", f"Hiba történt a terv duplikálása során: {e}")


# A plan_creator.py közvetlen elinditása
if __name__ == "__main__":
    root = tk.Tk()
    root.geometry("1920x1080")
    app = PlanCreatorPage(root)
    app.pack(expand=True, fill=tk.BOTH)
    root.mainloop()
