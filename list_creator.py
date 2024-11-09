import sqlite3
import tkinter as tk
from tkinter import Toplevel, PhotoImage, OptionMenu, StringVar, messagebox, Spinbox
from fpdf import FPDF
import openpyxl
from openpyxl.styles import Font
import os


class ListCreatorPage(tk.Frame):
    """A gyártási terv(ek) létrehozására szolgáló oldal."""

    def __init__(self, parent, *args, **kwargs):
        tk.Frame.__init__(self, parent, *args, **kwargs)
        self.configure(bg='white')

        # A termékek tárolására való tömb inicializálása
        self.product_entries = []

        # Az adatbázis kapcsolat
        self.conn = sqlite3.connect('cliptimizer.db')
        self.cursor = self.conn.cursor()

        # Hangers tábla ellenőrzése
        self.cursor.execute("SELECT COUNT(*) FROM hangers")
        hanger_count = self.cursor.fetchone()[0]
        if hanger_count == 0:
            # Ha üres, default érték
            self.cursor.execute("INSERT INTO hangers (total, available, occupied) VALUES (?, ?, ?)",
                                (70, 70, 0))
            self.conn.commit()

        # A függesztékek inicializálása
        self.available_hangers = 70
        self.occupied_hangers = 0

        # A függesztékek számainak lekérése
        self.update_hanger_status()

        # Products táblából adat lekérése
        self.cursor.execute("SELECT name FROM products")
        self.product_names = [row[0] for row in self.cursor.fetchall()]

        # Fő konténer a terv létrehozása és paneleknek és függesztékek számának
        main_container = tk.Frame(self, bg='white')
        main_container.pack(fill=tk.BOTH, expand=True)

        # Bal oldali konténer a terv paneleknek
        left_container = tk.Frame(main_container, bg='white')
        left_container.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Jobb oldali konténer a függeszték adatoknak
        right_container = tk.Frame(main_container, bg='white')
        right_container.pack(side=tk.RIGHT, padx=20, pady=20, anchor=tk.N)

        # A cím
        label = tk.Label(left_container, text="TERVEK LÉTREHOZÁSA", bg='white', font=('Helvetica', 20, 'bold'), padx=20,
                         pady=20)
        label.pack(anchor=tk.N)

        # Frame a függeszték adatoknak
        hanger_frame = tk.Frame(right_container, bg='white')
        hanger_frame.pack(anchor=tk.N)

        # Elérhető függesztékek
        self.available_label = tk.Label(hanger_frame, text=f"Elérhető függesztékek: {self.available_hangers}",
                                        bg='white', font=('Helvetica', 14))
        self.available_label.pack(anchor=tk.W, pady=10)

        # Elfoglalt függesztékek
        self.occupied_label = tk.Label(hanger_frame, text=f"Elfoglalt függesztékek: {self.occupied_hangers}",
                                       bg='white', font=('Helvetica', 14))
        self.occupied_label.pack(anchor=tk.W, pady=10)

        # Ezen az oldalon található képek
        self.add_icon = PhotoImage(file='images/add_plan_resized.png')
        self.eye_icon = PhotoImage(file='images/eye_resized.png')
        self.trash_icon = PhotoImage(file='images/trash_resized.png')
        self.print_icon = PhotoImage(file='images/print_resized.png')
        self.excel_icon = PhotoImage(file='images/file-excel_resized.png')
        self.duplicate_icon = PhotoImage(file='images/duplicate_resized.png')

        # Terv létrehozása gomb
        add_button = tk.Button(left_container, image=self.add_icon, bg='white', bd=0, command=self.create_new_plan)
        add_button.pack(anchor=tk.N, pady=20)

        # Terv panelek konténere
        self.plan_container = tk.Frame(left_container, bg='white')
        self.plan_container.pack(anchor=tk.N, pady=20)

        # Tervek betöltése
        self.load_plans()

    def update_hanger_status(self):
        """Lekérdezzük az elérhető és elfoglalt függesztékek számát."""
        self.cursor.execute("SELECT available, occupied FROM hangers")
        hanger_status = self.cursor.fetchone()

        if hanger_status:
            self.available_hangers, self.occupied_hangers = hanger_status
        else:
            # Ha nincs adat, default érték
            self.available_hangers, self.occupied_hangers = 70, 0

    def refresh_hanger_display(self):
        """Frissítjük a megjelenített függesztékek számát."""
        self.update_hanger_status()
        if self.available_hangers is not None and self.occupied_hangers is not None:
            print(
                f"Display update - Available hangers: {self.available_hangers}, Occupied hangers: "
                f"{self.occupied_hangers}")
            self.available_label.config(text=f"Elérhető függesztékek: {self.available_hangers}")
            self.occupied_label.config(text=f"Elfoglalt függesztékek: {self.occupied_hangers}")
        else:
            print("Display update failed: Missing hangers data.")
            self.available_label.config(text="Elérhető függesztékek: N/A")
            self.occupied_label.config(text="Elfoglalt függesztékek: N/A")

    def load_plans(self):
        """Tervek betöltése és panelek megjelenítése."""
        self.cursor.execute("SELECT plan_name, SUM(hangers_needed) FROM plans GROUP BY plan_name")
        plans = self.cursor.fetchall()

        # Minden tervhez létrehozunk egy panelt
        for plan in plans:
            plan_name, hangers_needed = plan
            self.add_plan_panel(plan_name, hangers_needed)

    def create_new_plan(self):
        """Egy terv létrehozása: Új ablak és adatbázis kapcsolódás."""
        new_window = Toplevel(self)
        new_window.title("Terv")
        new_window.geometry("1920x1080")

        # Ürítjük a product_entries listát, hogy ne legyen widget probléma
        self.product_entries.clear()

        # Fő konténer két oszlopban : bal oldalon a tartalom, jobb oldalon a függesztékek
        main_container = tk.Frame(new_window, bg='white')
        main_container.pack(fill=tk.BOTH, expand=True)

        left_content = tk.Frame(main_container, bg='white')
        left_content.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=20, pady=20)

        right_status_display = tk.Frame(main_container, bg='white')
        right_status_display.pack(side=tk.RIGHT, anchor=tk.N, padx=20, pady=20)

        # A cím
        label = tk.Label(left_content, text="Terv létrehozása", font=('Helvetica', 14))
        label.pack(pady=20)

        # A függesztékek
        available_label = tk.Label(right_status_display, text=f"Elérhető függesztékek: {self.available_hangers}",
                                   bg='white', font=('Helvetica', 14))
        available_label.pack(anchor=tk.W, pady=10)

        occupied_label = tk.Label(right_status_display, text=f"Elfoglalt függesztékek: {self.occupied_hangers}",
                                  bg='white', font=('Helvetica', 14))
        occupied_label.pack(anchor=tk.W, pady=10)

        # A terv neve
        plan_name_entry = tk.Entry(left_content)
        plan_name_entry.pack(pady=10)

        # Frame a termékeknek és az új termék gombnak
        product_frame = tk.Frame(left_content)
        product_frame.pack(expand=True, fill=tk.BOTH, pady=10)

        # Kezdő termék hozzáadása
        self.add_product_field(product_frame)

        # Az új termék hozzáadása gomb
        add_product_button = tk.Button(left_content, image=self.add_icon, bg='white', bd=0,
                                       command=lambda: self.add_product_field(product_frame))
        add_product_button.pack(pady=10)

        # Frame a mentés gombnak
        button_frame = tk.Frame(left_content)
        button_frame.pack(side=tk.BOTTOM, fill=tk.X, pady=10)

        # Mentés gomb hozzáadása
        save_button = tk.Button(button_frame, text="Mentés", font=('Helvetica', 14),
                                command=lambda: self.save_plan(new_window, plan_name_entry.get()))
        save_button.pack(pady=10)

    def add_product_field(self, window):
        """Termék, mennyiség és opcionális manuális függesztékszám mezők hozzáadása."""
        product_frame = tk.Frame(window)
        product_frame.pack(pady=10)

        # Legördülő menü a termék nevével
        selected_product = StringVar(window)
        selected_product.set(self.product_names[0])
        product_menu = OptionMenu(product_frame, selected_product, *self.product_names)
        product_menu.pack(side=tk.LEFT, padx=10)

        # Spinbox mennyiség mező
        amount_entry = Spinbox(product_frame, from_=1, to=1000, width=5)
        amount_entry.pack(side=tk.LEFT, padx=10)

        # Függesztékszám kijelző
        hanger_label = tk.Label(product_frame, text="Függesztékszám: 0", font=('Helvetica', 12))
        hanger_label.pack(side=tk.LEFT, padx=10)

        # Checkbox a manuális függesztékszám megadásához
        manual_hanger_var = tk.BooleanVar()
        manual_hanger_check = tk.Checkbutton(product_frame, text="Manuális függesztékszám",
                                             variable=manual_hanger_var, font=('Helvetica', 12))
        manual_hanger_check.pack(side=tk.LEFT, padx=10)

        # Beviteli mező a manuális függesztékszám megadásához (csak akkor aktív, ha be van pipálva)
        manual_hanger_entry = tk.Entry(product_frame, font=('Helvetica', 12), state='disabled', width=5)
        manual_hanger_entry.pack(side=tk.LEFT, padx=10)

        def toggle_hanger_entry():
            """Engedélyezi vagy tiltja a manuális függesztékszám mezőt a checkbox állapotától függően."""
            if manual_hanger_var.get():
                manual_hanger_entry.config(state='normal')
            else:
                manual_hanger_entry.config(state='disabled')

        # Ha a checkboxot bejelölik vagy kiveszik, aktiváljuk/deaktiváljuk a mezőt
        manual_hanger_check.config(command=toggle_hanger_entry)

        def update_hanger_label(*args):
            """Frissíti a szükséges függesztékszámot a megadott mennyiség alapján."""
            product_name = selected_product.get()
            amount = int(amount_entry.get())

            # Lekérjük a termék függeszték kapacitását
            self.cursor.execute("SELECT items_per_hanger FROM products WHERE name = ?", (product_name,))
            items_per_hanger = self.cursor.fetchone()[0]

            # Számoljuk ki a szükséges függesztékszámot
            required_hangers = (amount + items_per_hanger - 1) // items_per_hanger
            hanger_label.config(text=f"Függesztékszám: {required_hangers}")

        # Minden változáskor újraszámítja a szükséges függesztékszámot
        amount_entry.bind("<KeyRelease>", update_hanger_label)
        selected_product.trace("w", update_hanger_label)

        self.product_entries.append((selected_product, amount_entry, manual_hanger_var, manual_hanger_entry))

    def save_plan(self, window, plan_name):
        """Terv mentése az összes termékkel, mennyiséggel és a függesztékszámmal."""
        if not plan_name:
            messagebox.showerror("Hiba", "Terv neve nem lehet üres!")
            return

        # Ellenőrizzük, hogy létezik-e már ilyen nevű terv
        self.cursor.execute("SELECT COUNT(*) FROM plans WHERE plan_name = ?", (plan_name,))
        plan_exists = self.cursor.fetchone()[0]
        if plan_exists:
            messagebox.showerror("Hiba", "Már létezik ilyen nevű terv!")
            return

        # Ha vannak termékek, feldolgozzuk őket
        if self.product_entries:
            total_required_hangers = 0
            try:
                # Mentés adatgyűjtés
                products_to_save = []
                for selected_product, amount_entry, manual_hanger_var, manual_hanger_entry in self.product_entries:
                    product_name = selected_product.get()
                    amount = int(amount_entry.get())

                    # Product ID és items_per_hanger lekérése
                    self.cursor.execute("SELECT id, items_per_hanger FROM products WHERE name = ?",
                                        (product_name,))
                    product_data = self.cursor.fetchone()
                    product_id, items_per_hanger = product_data

                    # Függesztékszám számítása
                    required_hangers = (amount + items_per_hanger - 1) // items_per_hanger

                    if manual_hanger_var.get():
                        manual_hangers = int(manual_hanger_entry.get())
                        # Ellenőrizzük, hogy a manuálisan megadott függesztékszám kisebb-e, mint a szükséges
                        if manual_hangers < required_hangers:
                            messagebox.showerror("Hiba",
                                                 f"A megadott függesztékszám túl kevés! Legalább "
                                                 f"{required_hangers} függesztékre van szükség.")
                            return
                        required_hangers = manual_hangers

                    # Ellenőrzés, hogy van-e elég függeszték
                    if required_hangers > self.available_hangers:
                        messagebox.showerror("Hiba", "Nincs elég elérhető függeszték a tervhez.")
                        return

                    total_required_hangers += required_hangers
                    products_to_save.append((plan_name, product_id, amount, required_hangers))

                # Adatok mentése
                for plan_name, product_id, amount, required_hangers in products_to_save:
                    self.cursor.execute(
                        "INSERT INTO plans (plan_name, product_ID, amount, hangers_needed) "
                        "VALUES (?, ?, ?, ?)",
                        (plan_name, product_id, amount, required_hangers)
                    )
                    self.conn.commit()

                    # Függesztékek frissítése
                    self.cursor.execute("UPDATE hangers SET available = available - ?, occupied = occupied + ?",
                                        (required_hangers, required_hangers))
                    self.conn.commit()

                # Terv panel hozzáadása
                self.add_plan_panel(plan_name, total_required_hangers)

                self.refresh_hanger_display()

                # Zárjuk be az ablakot
                window.destroy()

            except Exception as e:
                messagebox.showerror("Hiba", f"Hiba történt a terv mentésekor: {e}")
                return

    def add_plan_panel(self, plan_name, hangers_needed):
        """Panel hozzáadása a tervhez és teljes ciklusidő formázása."""
        # Teljes ciklusidő lekérdezése a tervhez
        self.cursor.execute("""
            SELECT p.total_cycle_time, pl.hangers_needed
            FROM plans pl
            JOIN products p ON pl.product_ID = p.id
            WHERE pl.plan_name = ?
        """, (plan_name,))
        plans = self.cursor.fetchall()

        # Teljes ciklusidő
        total_cycle_time = sum(hangers_needed * cycle_time for cycle_time, hangers_needed in plans)

        # Ciklusidő konvertálása nap, óra, perc, másodperc formátumba
        days, remainder = divmod(total_cycle_time, 86400)
        hours, remainder = divmod(remainder, 3600)
        minutes, seconds = divmod(remainder, 60)
        formatted_cycle_time = f"{days}n:{hours}ó:{minutes}p:{seconds}mp"

        panel = tk.Frame(self.plan_container, bg='lightgrey', bd=2, relief='solid')
        panel.pack(pady=5, padx=10, fill=tk.X)

        # Terv neve és attribútumok
        plan_label = tk.Label(panel, text=f"{plan_name}", bg='lightgrey', font=('Helvetica', 12))
        plan_label.pack(side=tk.LEFT, padx=10, pady=5)

        # Elfoglalt függesztékek és ciklusidő megjelenítése
        hangers_label = tk.Label(panel,
                                 text=f"Elfoglalt függesztékek: {hangers_needed}, Ciklusidő: {formatted_cycle_time}",
                                 bg='lightgrey', font=('Helvetica', 10))
        hangers_label.pack(side=tk.LEFT, padx=10, pady=5)

        # Terv megnézése, törlése, nyomtatása és exportálása
        view_button = tk.Button(panel, image=self.eye_icon, bg='lightgrey', bd=0,
                                command=lambda: self.view_plan(plan_name))
        view_button.pack(side=tk.RIGHT, padx=10, pady=5)

        delete_button = tk.Button(panel, image=self.trash_icon, bg='lightgrey', bd=0,
                                  command=lambda: self.confirm_delete(plan_name))
        delete_button.pack(side=tk.RIGHT, padx=10, pady=5)

        print_button = tk.Button(panel, image=self.print_icon, bg='lightgrey', bd=0,
                                 command=lambda: self.print_plan(plan_name))
        print_button.pack(side=tk.RIGHT, padx=10, pady=5)

        excel_button = tk.Button(panel, image=self.excel_icon, bg='lightgrey', bd=0,
                                 command=lambda: self.export_plan_to_excel(plan_name))
        excel_button.pack(side=tk.RIGHT, padx=10, pady=5)

        duplicate_button = tk.Button(panel, image=self.duplicate_icon, bg='lightgrey', bd=0,
                                     command=lambda: self.duplicate_plan(plan_name))
        duplicate_button.pack(side=tk.RIGHT, padx=10, pady=5)

    def view_plan(self, plan_name):
        """Terv megjelenítése egy új ablakban."""
        new_window = Toplevel(self)
        new_window.title(f"Terv: {plan_name}")
        new_window.geometry("1920x1080")

        label = tk.Label(new_window, text=f"Terv neve: {plan_name}", font=('Helvetica', 12, 'bold'))
        label.pack(pady=10)

        # A termékek megjelenítése a tervhez
        self.cursor.execute("SELECT product_ID, amount, hangers_needed FROM plans WHERE plan_name = ?",
                            (plan_name,))
        products = self.cursor.fetchall()

        for product_id, amount, hangers_needed in products:
            # Lekérjük az összes szükséges attribútumot
            self.cursor.execute(
                "SELECT name, color, clip_type, items_per_hanger, total_cycle_time, material_per_part "
                "FROM products WHERE id = ?",
                (product_id,)
            )
            product = self.cursor.fetchone()
            product_name, color, clip_type, items_per_hanger, total_cycle_time, material_per_part = product

            # Képletek az összes anyagszükséglet és a függesztékek ciklusidejéhez
            total_material = amount * material_per_part
            total_cycle_for_hangers = hangers_needed * total_cycle_time

            # Frame a termékhez tartozó attribútumok megjelenítésére
            frame = tk.Frame(new_window, bg='white', bd=1, relief='solid')
            frame.pack(pady=5, padx=10, fill=tk.X)

            attributes = [
                f"Termék neve: {product_name}",
                f"Mennyiség: {amount}",
                f"Szín: {color}",
                f"Klipsz típusa: {clip_type}",
                f"Függesztékre rakható: {items_per_hanger} db",
                f"Függesztékenként ciklusidő: {total_cycle_time} mp",
                f"Lefoglalt függesztékszám: {hangers_needed}",
                f"Ennyi függesztékre teljes ciklusidő: {total_cycle_for_hangers} mp",
                f"Anyagszükséglet / alkatrész: {material_per_part:.2f} g",
                f"Összes anyagszükséglet: {total_material:.2f} g"
            ]

            # Attribútumok hozzáadása a frame-hez
            for attr in attributes:
                label = tk.Label(frame, text=attr, font=('Helvetica', 10), bg='white')
                label.pack(side=tk.LEFT, padx=5, pady=5)

    def confirm_delete(self, plan_name):
        """Terv törlésének megerősítése és törlése."""
        if messagebox.askyesno("Törlés megerősítése", f"Biztos törölni szeretné ezt a tervet : "
                                                      f"'{plan_name}'?"):
            self.delete_plan(plan_name)

    def delete_plan(self, plan_name):
        """Terv törlése az adatbázisból és a panelből."""
        try:
            # Tervhez tartozó foglalt függesztékek száma
            self.cursor.execute("SELECT SUM(hangers_needed) FROM plans WHERE plan_name = ?",
                                (plan_name,))
            hangers_to_free = self.cursor.fetchone()[0]

            if hangers_to_free:
                # Frissítjük a függesztékek számát
                self.cursor.execute("UPDATE hangers SET available = available + ?, occupied = occupied - ?",
                                    (hangers_to_free, hangers_to_free))
                self.conn.commit()

            # Töröljük a tervet
            self.cursor.execute("DELETE FROM plans WHERE plan_name = ?", (plan_name,))
            self.conn.commit()

            # Frissítjük a felhasználói felületet
            for widget in self.plan_container.winfo_children():
                if isinstance(widget, tk.Frame):
                    if widget.winfo_children()[0].cget("text") == plan_name:
                        widget.destroy()

            # Függesztékek frissítése
            self.refresh_hanger_display()

        except Exception as e:
            messagebox.showerror("Hiba", f"Hiba történt a terv törlésekor: {e}")

    def print_plan(self, plan_name):
        """Terv nyomtatása PDF formátumban."""
        try:
            pdf = FPDF(orientation='L', unit='mm', format='A4')
            pdf.add_page()

            # Egyedi font
            pdf.add_font('DejaVu', '', 'fonts/DejaVuSans.ttf', uni=True)
            pdf.add_font('DejaVu', 'B', 'fonts/DejaVuSans-Bold.ttf', uni=True)

            # Táblázat attribútumok
            cell_height = 6
            font_size = 8
            pdf.set_font('DejaVu', 'B', font_size)

            # A terv neve
            pdf.cell(0, 8, f"TERV: {plan_name}", ln=True, align='C')

            # A headerek
            headers = ["Kód", "Mennyiség", "Szín", "Klipsz", "Füg-re rakható",
                       "Füg-ként ciklusidő", "Lefoglalt füg.", "Anyagszükséglet (g)",
                       "Ennyi füg-re teljes idő"]
            col_widths = [30, 20, 20, 20, 40, 30, 35, 40, 35]
            for i, header in enumerate(headers):
                pdf.cell(col_widths[i], cell_height, header, border=1, align='C')
            pdf.ln(cell_height)

            # A termékek
            pdf.set_font('DejaVu', '', font_size)
            self.cursor.execute("SELECT product_ID, amount, hangers_needed FROM plans WHERE plan_name = ?",
                                (plan_name,))
            products = self.cursor.fetchall()

            for product_id, amount, hangers_needed in products:
                self.cursor.execute(
                    "SELECT name, color, clip_type, items_per_hanger, total_cycle_time, material_per_part FROM "
                    "products WHERE id = ?",
                    (product_id,)
                )
                product = self.cursor.fetchone()
                product_name, color, clip_type, items_per_hanger, total_cycle_time, material_per_part = product

                # Számítások
                total_material = round(material_per_part * amount, 2)
                total_cycle_for_hangers = hangers_needed * total_cycle_time

                # Sor az adatokkal
                row = [
                    product_name, str(amount), color, clip_type, str(items_per_hanger),
                    str(total_cycle_time), str(hangers_needed), f"{total_material} g", f"{total_cycle_for_hangers} mp"
                ]

                for i, item in enumerate(row):
                    pdf.cell(col_widths[i], cell_height, item, border=1, align='C')
                pdf.ln(cell_height)

            # A PDF mentése és megnyitása
            pdf_output_path = f"{plan_name}_terv.pdf"
            pdf.output(pdf_output_path)

            try:
                os.startfile(pdf_output_path, "print")
            except FileNotFoundError:
                messagebox.showerror("Hiba", "Nincs alapértelmezett alkalmazás PDF-hez. "
                                             "Telepítsen egy PDF-olvasót.")

            messagebox.showinfo("Nyomtatás", "A terv nyomtatása folyamatban.")

        except Exception as e:
            messagebox.showerror("Hiba", f"Hiba történt a terv nyomtatása során: {e}")

    def export_plan_to_excel(self, plan_name):
        """Terv exportálása Excel formátumban."""
        try:
            # Excel létrehozása
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = plan_name

            bold_font = Font(bold=True)
            title_font = Font(size=14, bold=True)

            # Terv neve
            ws.merge_cells("A1:J1")
            ws["A1"] = plan_name.upper()
            ws["A1"].font = title_font

            # Adatfejlécek és új oszlop a teljes anyagszükséglet számára
            headers = ["Termék neve", "Mennyiség", "Szín", "Klipsz típusa", "Függesztékre helyezhető darabszám",
                       "Függesztékenként ciklusidő (sec)", "Lefoglalt függesztékek száma", "Összes anyagszükséglet (g)",
                       "Ennyi függesztékre teljes ciklusidő (sec)"]
            ws.append([""] * len(headers))
            ws.append(headers)
            for col in range(1, len(headers) + 1):
                ws.cell(row=3, column=col).font = bold_font

            # Termékek megjelenítése a tervhez
            self.cursor.execute("SELECT product_ID, amount, hangers_needed FROM plans WHERE plan_name = ?",
                                (plan_name,))
            products = self.cursor.fetchall()

            for product_id, amount, reserved_hangers in products:
                self.cursor.execute(
                    "SELECT name, color, clip_type, items_per_hanger, total_cycle_time, material_per_part FROM "
                    "products WHERE id = ?",
                    (product_id,)
                )
                product = self.cursor.fetchone()
                product_name, color, clip_type, items_per_hanger, total_cycle_time, material_per_part = product

                # Anyagszükséglet és teljes ciklusidő számítása
                total_material = amount * material_per_part
                total_cycle_for_hangers = reserved_hangers * total_cycle_time

                # Adatok hozzáadása Excelhez
                ws.append([product_name, amount, color, clip_type, items_per_hanger, total_cycle_time, reserved_hangers,
                           f"{total_material:.2f}", total_cycle_for_hangers])

            # Excel mentése
            excel_output_path = f"{plan_name}_terv.xlsx"
            wb.save(excel_output_path)

            # Excel megnyitása
            os.startfile(excel_output_path)

            messagebox.showinfo("Exportálás", "A terv sikeresen exportálva Excel formátumba.")

        except Exception as e:
            messagebox.showerror("Hiba", f"Hiba történt a terv Excel exportálása során: {e}")

    def duplicate_plan(self, plan_name):
        """Egy terv duplikálása."""
        try:
            # Lekérjük a tervhez tartozó termékeket és azok adatait
            self.cursor.execute("SELECT product_ID, amount FROM plans WHERE plan_name = ?",
                                (plan_name,))
            products = self.cursor.fetchall()

            # Új ablak megnyitása a duplikált terv szerkesztéséhez
            new_window = Toplevel(self)
            new_window.title(f"Duplikált terv: {plan_name}")
            new_window.geometry("1920x1080")

            # Új terv neve (opcionálisan megváltoztatható)
            new_plan_name_entry = tk.Entry(new_window)
            new_plan_name_entry.insert(0, f"{plan_name}_másolat")
            new_plan_name_entry.pack(pady=10)

            # Termékek és mennyiség mezők inicializálása
            product_frame = tk.Frame(new_window)
            product_frame.pack(expand=True, fill=tk.BOTH, pady=10)

            # Termékek hozzáadása a duplikált tervhez
            self.product_entries = []  # Ürítjük a meglévő termékbejegyzéseket
            for product_id, amount in products:
                # Lekérjük a termék részleteit
                self.cursor.execute(
                    "SELECT name FROM products WHERE id = ?",
                    (product_id,)
                )
                product_name = self.cursor.fetchone()[0]

                # Termék mezők megjelenítése
                self.add_product_field(product_frame)
                self.product_entries[-1][0].set(product_name)  # Termék neve
                self.product_entries[-1][1].delete(0, 'end')
                self.product_entries[-1][1].insert(0, amount)  # Mennyiség

            # + ikon hozzáadása a termékekhez
            add_product_button = tk.Button(new_window, image=self.add_icon, bg='white', bd=0,
                                           command=lambda: self.add_product_field(product_frame))
            add_product_button.pack(pady=10)

            # Mentés gomb hozzáadása a duplikált terv mentéséhez
            save_button = tk.Button(new_window, text="Mentés", font=('Helvetica', 14),
                                    command=lambda: self.save_plan(new_window, new_plan_name_entry.get()))
            save_button.pack(pady=10)

        except Exception as e:
            messagebox.showerror("Hiba", f"Hiba történt a terv duplikálása során: {e}")


# A list_creator.py közvetlen elinditása
if __name__ == "__main__":
    root = tk.Tk()
    root.geometry("1920x1080")
    app = ListCreatorPage(root)
    app.pack(expand=True, fill=tk.BOTH)
    root.mainloop()
